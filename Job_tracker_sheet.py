import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import FormulaRule
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime
import os

def create_excel_job_tracker(output_file='Job_Tracker.xlsx'):
    """Create an Excel job tracker with dashboard"""
    
    # Define column headers for main tracker
    headers = [
        '#', 'Company Name', 'Job Title', 'Job Location', 'Date Applied',
        'Job Posting Link', 'Resume Link', 'Cover Letter Link', 
        'Application Status', 'Follow-Up Date', 'Response Received?', 'Notes'
    ]
    
    # Create sample data for demonstration
    sample_data = [
        {
            '#': 1,
            'Company Name': 'Tech Corp',
            'Job Title': 'Software Engineer',
            'Date Applied': datetime.today().strftime('%Y-%m-%d'),
            'Application Status': 'Applied'
        },
        {
            '#': 2,
            'Company Name': 'Data Inc',
            'Job Title': 'Data Analyst',
            'Date Applied': (datetime.today() - pd.Timedelta(days=3)).strftime('%Y-%m-%d'),
            'Application Status': 'Interview Scheduled'
        }
    ]
    
    # Create DataFrame
    df = pd.DataFrame(columns=headers)
    df = pd.concat([df, pd.DataFrame(sample_data)], ignore_index=True)
    
    # Create Excel writer
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Job Applications')
    
    # Get workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Job Applications']
    
    # Format headers
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
    
    # Set column widths
    column_widths = {
        'A': 5,    # #
        'B': 25,   # Company Name
        'C': 25,   # Job Title
        'D': 20,   # Job Location
        'E': 15,   # Date Applied
        'F': 40,   # Job Posting Link
        'G': 40,   # Resume Link
        'H': 40,   # Cover Letter Link
        'I': 20,   # Application Status
        'J': 15,   # Follow-Up Date
        'K': 20,   # Response Received?
        'L': 40    # Notes
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width
    
    # Add data validation for Application Status
    dv = DataValidation(type="list", formula1='"Applied,Interview Scheduled,Offer,Rejected,Followed Up,No Response,On Hold"', 
                       allow_blank=True)
    worksheet.add_data_validation(dv)
    dv.add(f'I2:I1048576')  # Apply to entire column except header
    
    # Define conditional formatting colors
    status_colors = {
        'Applied': 'ADD8E6',          # Light Blue
        'Interview Scheduled': '90EE90',  # Light Green
        'Offer': 'FFFF00',            # Gold/Yellow
        'Rejected': 'FF9999',          # Light Red
        'Followed Up': 'FFA500',       # Orange
        'No Response': 'D3D3D3',       # Light Gray
        'On Hold': 'E6E6FA'           # Lavender
    }
    
    # Apply conditional formatting
    for status, color in status_colors.items():
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        formula = f'$I2="{status}"'
        
        # Create a Rule object
        rule = FormulaRule(formula=[formula], fill=fill)
        
        # Apply to entire row (columns A-L)
        for col in range(1, 13):
            column_letter = get_column_letter(col)
            worksheet.conditional_formatting.add(
                f'{column_letter}2:{column_letter}1048576',
                rule
            )
    
    # Freeze top row
    worksheet.freeze_panes = 'A2'
    
    # ========== DASHBOARD CREATION ==========
    dashboard = workbook.create_sheet(title="📈 Dashboard")
    
    # Dashboard layout
    dashboard.merge_cells('A1:D1')
    dashboard['A1'] = "Job Application Dashboard"
    dashboard['A1'].font = Font(size=18, bold=True)
    dashboard['A1'].alignment = Alignment(horizontal='center')
    
    # Create metric cards
    metrics = [
        ("Total Applied", "COUNTIF('Job Applications'!I:I,\"*\")"),
        ("Interviews", "COUNTIF('Job Applications'!I:I,\"Interview Scheduled\")"),
        ("Offers", "COUNTIF('Job Applications'!I:I,\"Offer\")"),
        ("Rejections", "COUNTIF('Job Applications'!I:I,\"Rejected\")"),
        ("Follow-ups Needed", "COUNTIF('Job Applications'!I:I,\"Followed Up\")"),
        ("No Response", "COUNTIF('Job Applications'!I:I,\"No Response\")")
    ]
    
    # Position metric cards
    for i, (label, formula) in enumerate(metrics):
        row = 3 + (i // 2) * 3
        col = 2 + (i % 2) * 5
        
        # Metric label
        dashboard.cell(row=row, column=col, value=label).font = Font(bold=True)
        
        # Metric value
        dashboard.cell(row=row+1, column=col, value=f'={formula}').font = Font(size=14, bold=True)
        
        # Format as card
        for r in range(row, row+2):
            for c in range(col, col+3):
                cell = dashboard.cell(row=r, column=c)
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Create pie chart for status distribution
    pie = PieChart()
    labels = Reference(worksheet, min_col=9, min_row=2, max_row=100)
    data = Reference(worksheet, min_col=9, min_row=1, max_row=100)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Application Status Distribution"
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    dashboard.add_chart(pie, "A15")
    
    # Create bar chart for daily applications
    bar = BarChart()
    dates = Reference(worksheet, min_col=5, min_row=2, max_row=100)
    counts = Reference(worksheet, min_col=1, min_row=1, max_row=100)
    bar.add_data(counts, titles_from_data=True)
    bar.set_categories(dates)
    bar.title = "Daily Applications"
    bar.style = 10
    bar.y_axis.title = "Count"
    bar.x_axis.title = "Date"
    dashboard.add_chart(bar, "I15")
    
    # Save the workbook
    workbook.save(output_file)
    print(f"Excel job tracker with dashboard created successfully: {output_file}")

def create_google_sheets_job_tracker(creds_file=None, sheet_name='Job Tracker'):
    """Create a Google Sheets job tracker with dashboard"""
    
    if not creds_file or not os.path.exists(creds_file):
        print("Google Sheets creation requires a credentials file.")
        return
    
    try:
        # Authenticate with Google Sheets API
        scopes = ['https://www.googleapis.com/auth/spreadsheets',
                 'https://www.googleapis.com/auth/drive']
        
        creds = Credentials.from_service_account_file(creds_file, scopes=scopes)
        client = gspread.authorize(creds)
        
        # Create a new spreadsheet
        spreadsheet = client.create(sheet_name)
        
        # Get the first worksheet (main tracker)
        worksheet = spreadsheet.get_worksheet(0)
        worksheet.update_title("Job Applications")
        
        # Define column headers
        headers = [
            '#', 'Company Name', 'Job Title', 'Job Location', 'Date Applied',
            'Job Posting Link', 'Resume Link', 'Cover Letter Link', 
            'Application Status', 'Follow-Up Date', 'Response Received?', 'Notes'
        ]
        
        # Update the worksheet with headers
        worksheet.append_row(headers)
        
        # Format headers (bold)
        header_format = {
            "textFormat": {"bold": True}
        }
        worksheet.format('A1:L1', header_format)
        
        # Set column widths
        column_widths = {
            1: 50,    # #
            2: 150,   # Company Name
            3: 150,   # Job Title
            4: 120,   # Job Location
            5: 100,   # Date Applied
            6: 250,   # Job Posting Link
            7: 250,   # Resume Link
            8: 250,   # Cover Letter Link
            9: 120,   # Application Status
            10: 100,   # Follow-Up Date
            11: 150,   # Response Received?
            12: 250    # Notes
        }
        
        # Google Sheets uses pixel widths
        worksheet.resize_columns(column_widths)
        
        # Add data validation for Application Status
        validation_rule = {
            "condition": {
                "type": "ONE_OF_LIST",
                "values": ["Applied", "Interview Scheduled", "Offer", "Rejected", 
                          "Followed Up", "No Response", "On Hold"]
            },
            "strict": True,
            "showCustomUi": True
        }
        
        # Apply to column I (9th column)
        start_cell = gspread.utils.rowcol_to_a1(2, 9)
        end_cell = gspread.utils.rowcol_to_a1(1000, 9)
        range_notation = f"{start_cell}:{end_cell}"
        
        worksheet.set_data_validation(range_notation, validation_rule)
        
        # Define conditional formatting rules
        status_colors = {
            'Applied': {'red': 0.678, 'green': 0.847, 'blue': 0.902},
            'Interview Scheduled': {'red': 0.565, 'green': 0.933, 'blue': 0.565},
            'Offer': {'red': 1.0, 'green': 1.0, 'blue': 0.0},
            'Rejected': {'red': 1.0, 'green': 0.6, 'blue': 0.6},
            'Followed Up': {'red': 1.0, 'green': 0.647, 'blue': 0.0},
            'No Response': {'red': 0.827, 'green': 0.827, 'blue': 0.827},
            'On Hold': {'red': 0.902, 'green': 0.902, 'blue': 0.980}
        }
        
        # Apply conditional formatting
        requests = []
        for status, color in status_colors.items():
            rule = {
                "ranges": [{
                    "sheetId": worksheet.id,
                    "startRowIndex": 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": 12
                }],
                "booleanRule": {
                    "condition": {
                        "type": "CUSTOM_FORMULA",
                        "values": [{"userEnteredValue": f'=$I2="{status}"'}]
                    },
                    "format": {
                        "backgroundColor": color
                    }
                }
            }
            requests.append({"addConditionalFormatRule": {
                "index": 0,
                "rule": rule
            }})
        
        # Freeze the top row
        freeze_request = {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": worksheet.id,
                    "gridProperties": {
                        "frozenRowCount": 1
                    }
                },
                "fields": "gridProperties.frozenRowCount"
            }
        }
        requests.append(freeze_request)
        
        # ========== CREATE DASHBOARD ==========
        dashboard = spreadsheet.add_worksheet(title="📈 Dashboard", rows=100, cols=20)
        
        # Dashboard header
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": dashboard.id,
                    "startRowIndex": 0,
                    "endRowIndex": 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": 4
                },
                "cell": {
                    "userEnteredValue": {"stringValue": "Job Application Dashboard"},
                    "userEnteredFormat": {
                        "textFormat": {"fontSize": 18, "bold": True},
                        "horizontalAlignment": "CENTER"
                    }
                },
                "fields": "userEnteredValue,userEnteredFormat.textFormat,userEnteredFormat.horizontalAlignment"
            },
            "mergeCells": {
                "range": {
                    "sheetId": dashboard.id,
                    "startRowIndex": 0,
                    "endRowIndex": 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": 4
                },
                "mergeType": "MERGE_ALL"
            }
        })
        
        # Add metric cards
        metrics = [
            ("A3", "Total Applied", "=COUNTIF('Job Applications'!I:I,\"*\")"),
            ("G3", "Interviews", "=COUNTIF('Job Applications'!I:I,\"Interview Scheduled\")"),
            ("A8", "Offers", "=COUNTIF('Job Applications'!I:I,\"Offer\")"),
            ("G8", "Rejections", "=COUNTIF('Job Applications'!I:I,\"Rejected\")"),
            ("A13", "Follow-ups Needed", "=COUNTIF('Job Applications'!I:I,\"Followed Up\")"),
            ("G13", "No Response", "=COUNTIF('Job Applications'!I:I,\"No Response\")")
        ]
        
        for cell, label, formula in metrics:
            # Label
            requests.append({
                "updateCells": {
                    "range": {
                        "sheetId": dashboard.id,
                        "startRowIndex": int(cell[1:])-1,
                        "endRowIndex": int(cell[1:]),
                        "startColumnIndex": ord(cell[0].upper())-65,
                        "endColumnIndex": ord(cell[0].upper())-64
                    },
                    "rows": [{
                        "values": [{
                            "userEnteredValue": {"stringValue": label},
                            "userEnteredFormat": {"textFormat": {"bold": True}}
                        }]
                    }],
                    "fields": "userEnteredValue,userEnteredFormat.textFormat"
                }
            })
            
            # Value
            requests.append({
                "updateCells": {
                    "range": {
                        "sheetId": dashboard.id,
                        "startRowIndex": int(cell[1:]),
                        "endRowIndex": int(cell[1:])+1,
                        "startColumnIndex": ord(cell[0].upper())-65,
                        "endColumnIndex": ord(cell[0].upper())-64
                    },
                    "rows": [{
                        "values": [{
                            "userEnteredValue": {"formulaValue": formula},
                            "userEnteredFormat": {
                                "textFormat": {"fontSize": 14, "bold": True},
                                "backgroundColor": {"red": 0.94, "green": 0.94, "blue": 0.94}
                            }
                        }]
                    }],
                    "fields": "userEnteredValue,userEnteredFormat"
                }
            })
        
        # Add charts
        pie_chart_request = {
            "addChart": {
                "chart": {
                    "spec": {
                        "title": "Application Status Distribution",
                        "pieChart": {
                            "legendPosition": "RIGHT_LEGEND",
                            "domain": {
                                "sourceRange": {
                                    "sources": [{
                                        "sheetId": worksheet.id,
                                        "startRowIndex": 1,
                                        "endRowIndex": 100,
                                        "startColumnIndex": 8,
                                        "endColumnIndex": 9
                                    }]
                                }
                            },
                            "series": {
                                "sourceRange": {
                                    "sources": [{
                                        "sheetId": worksheet.id,
                                        "startRowIndex": 0,
                                        "endRowIndex": 100,
                                        "startColumnIndex": 8,
                                        "endColumnIndex": 9
                                    }]
                                }
                            },
                            "pieHole": 0.4
                        }
                    },
                    "position": {
                        "overlayPosition": {
                            "anchorCell": {
                                "sheetId": dashboard.id,
                                "rowIndex": 15,
                                "columnIndex": 0
                            }
                        }
                    }
                }
            }
        }
        requests.append(pie_chart_request)
        
        bar_chart_request = {
            "addChart": {
                "chart": {
                    "spec": {
                        "title": "Daily Applications",
                        "barChart": {
                            "legendPosition": "BOTTOM_LEGEND",
                            "axis": [
                                {"title": "Date"},
                                {"title": "Count"}
                            ],
                            "domains": [{
                                "domain": {
                                    "sourceRange": {
                                        "sources": [{
                                            "sheetId": worksheet.id,
                                            "startRowIndex": 1,
                                            "endRowIndex": 100,
                                            "startColumnIndex": 4,
                                            "endColumnIndex": 5
                                        }]
                                    }
                                }
                            }],
                            "series": [{
                                "series": {
                                    "sourceRange": {
                                        "sources": [{
                                            "sheetId": worksheet.id,
                                            "startRowIndex": 0,
                                            "endRowIndex": 100,
                                            "startColumnIndex": 0,
                                            "endColumnIndex": 1
                                        }]
                                    }
                                },
                                "targetAxis": "LEFT_AXIS"
                            }]
                        }
                    },
                    "position": {
                        "overlayPosition": {
                            "anchorCell": {
                                "sheetId": dashboard.id,
                                "rowIndex": 15,
                                "columnIndex": 8
                            }
                        }
                    }
                }
            }
        }
        requests.append(bar_chart_request)
        
        # Batch update the spreadsheet
        spreadsheet.batch_update(requests)
        
        print(f"Google Sheets job tracker with dashboard created successfully: {spreadsheet.url}")
        return spreadsheet.url
        
    except Exception as e:
        print(f"Error creating Google Sheets job tracker: {str(e)}")
        return None

def main():
    print("Enhanced Job Tracker Spreadsheet Creator")
    print("1. Create Excel file (.xlsx) with Dashboard")
    print("2. Create Google Sheets with Dashboard (requires service account credentials)")
    
    choice = input("Enter your choice (1 or 2): ")
    
    if choice == '1':
        output_file = input("Enter output filename (default: Job_Tracker.xlsx): ") or 'Job_Tracker.xlsx'
        create_excel_job_tracker(output_file)
    elif choice == '2':
        creds_file = input("Enter path to your Google service account credentials JSON file: ")
        sheet_name = input("Enter name for your Google Sheet (default: Job Tracker): ") or 'Job Tracker'
        create_google_sheets_job_tracker(creds_file, sheet_name)
    else:
        print("Invalid choice. Please enter 1 or 2.")

if __name__ == "__main__":
    main()